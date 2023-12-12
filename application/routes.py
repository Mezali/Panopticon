import json

import openpyxl
import requests
import urllib3
from flask import render_template, session, flash, redirect, url_for, request, jsonify

from application import app, mongo
from application.forms import RegisterForm, LoginForm, RegisterColaborador, EditColaborador
from application.functions import fetchbravas, insertBravas, editkit, seluser, editbravas, cadMassa, delMassa, \
    delbravas, dellGeral

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
ip = app.config['BRAVAS_IP']


@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('cad_massa'))
    else:
        return redirect(url_for('login'))


@app.route('/login', methods=['POST', 'GET'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        user = mongo.db.user.find_one({'name': username})

        if user is not None:  # Check if the user exists
            if username == user['name'] and password == user['password']:
                session['username'] = user['name']
                return redirect(url_for('listar'))
            else:
                flash('Nome ou senha incorretos!', 'danger')

    return render_template('login.html', form=form)


@app.route('/register', methods=['POST', 'GET'])
def register():
    form = RegisterForm()

    if form.validate_on_submit():  # Checar a porra do login
        user = mongo.db.user.find_one({'name': f'{form.username.data}'})

        if user is None:  # Se não houver usuário cadastrado ele vai cadastrar
            mongo.db.user.insert_one(
                {'name': f'{form.username.data}', 'password': f'{form.password.data}'})
            session['username'] = form.username.data

            return redirect(url_for('index'))
        else:
            flash('Usuário já cadastrado!', 'danger')

    return render_template('register.html', form=form)


@app.route('/logout')
def logout():
    session.clear()  # Limpa a sessão do usuário
    flash('Você saiu com sucesso!', 'success')
    return redirect(url_for('login'))


@app.route('/cad-colaborador', methods=['POST', 'GET'])
def cad_colaborador():
    if 'username' in session:
        form = RegisterColaborador()
        if form.validate_on_submit():
            nome = form.nome.data
            matricula = form.matricula.data
            cartao = form.cartao.data
            seg = form.seg.data
            ter = form.ter.data
            qua = form.qua.data
            qui = form.qui.data
            sex = form.sex.data
            sab = form.sab.data
            dom = form.dom.data
            cafe_manha = form.cafe_manha.data
            almoco = form.almoco.data
            cafe_pendura = form.cafe_pendura.data
            cafe_tarde = form.cafe_tarde.data
            janta = form.janta.data

            try:
                response = insertBravas(ip, nome, matricula, cartao, seg, ter, qua, qui, sex, sab, dom, cafe_manha,
                                        almoco,
                                        cafe_pendura,
                                        cafe_tarde,
                                        janta)
            except requests.exceptions.RequestException as e:
                message = f"Erro na solicitação: {e}"
                return render_template('error.html', message=message), 500

            if response.status_code == 200:
                flash('Colaborador registrado com sucesso!', 'success')
            else:
                flash('Erro ao registrar colaborador! Verifique se o mesmo já foi registrado', 'danger')

        return render_template('cad-colaborador.html', form=form)
    else:
        return redirect(url_for('login'))


@app.route('/op-massa', methods=['POST', 'GET'])
def cad_massa():
    if 'username' in session:
        return render_template('op-massa.html')
    else:
        return redirect(url_for('login'))


@app.route('/massa-add', methods=['POST', 'GET'])
def massa_add():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo encontrado'})

        file = request.files['file']

        # Verifica se o arquivo tem um nome
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo inválido'})

        # Carrega o arquivo usando o Openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            matricula = row[1]
            tag = row[2]
            ativado = row[3]
            cafe_manha = row[4]
            almoco = row[5]
            cafe_pendura = row[6]
            cafe_tarde = row[7]
            janta = row[8]

            # Adiciona verificação para células vazias
            if nome is None:
                break

            print(f'{nome} - {matricula} - {tag}')

            cadMassa(ip=ip, nome=nome, matricula=matricula, tag=tag, ativado=ativado, cafe_manha=cafe_manha,
                     almoco=almoco,
                     cafe_pendura=cafe_pendura, cafe_tarde=cafe_tarde, janta=janta)

        # Retorna a resposta em JSON
        return jsonify({'status': 200, 'mensagem': 'Arquivo processado com sucesso!'})
    except Exception as e:
        print(f'Erro ao processar o arquivo: {str(e)}')
        return jsonify({'status': 500, 'mensagem': f'Erro ao processar o arquivo: {str(e)}'})


@app.route('/massa-del', methods=['POST', 'GET'])
def massa_del():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo encontrado'})

        file = request.files['file']

        # Verifica se o arquivo tem um nome
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo inválido'})

        # Carrega o arquivo usando o Openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            matricula = row[1]
            delMassa(ip=ip, name=nome, matricula=matricula)
            if nome is None:
                break

            print(f'{nome}')

        # Retorna a resposta em JSON
        return jsonify({'status': 200, 'mensagem': 'Arquivo processado com sucesso!'})
    except Exception as e:
        print(f'Erro ao processar o arquivo: {str(e)}')
        return jsonify({'status': 500, 'mensagem': f'Erro ao processar o arquivo: {str(e)}'})


@app.route('/massa-edit', methods=['POST', 'GET'])
def edit_massa():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo encontrado'})

        file = request.files['file']

        # Verifica se o arquivo tem um nome
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo inválido'})

        # Carrega o arquivo usando o Openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            matricula = row[1]
            tag = row[2]
            ativado = row[3]
            cafe_manha = row[4]
            almoco = row[5]
            cafe_pendura = row[6]
            cafe_tarde = row[7]
            janta = row[8]
            nome = f'{nome} - {matricula}'
            editbravas(ip=ip, nome=nome, tag=tag, ativado=ativado, cafe_manha=cafe_manha, almoco=almoco,
                       cafe_pendura=cafe_pendura, cafe_tarde=cafe_tarde, janta=janta)

        return jsonify({'status': 200, 'mensagem': 'Arquivo processado com sucesso!'})
    except Exception as e:
        print(f'Erro ao processar o arquivo: {str(e)}')
        return jsonify({'status': 500, 'mensagem': f'Erro ao processar o arquivo: {str(e)}'})


@app.route('/list')
def listar():
    if 'username' in session:
        try:
            response = fetchbravas(ip)

            if response.status_code == 200:
                data = json.loads(response.text)
                user = data["config"]["users"]

                return render_template("list-colaborador.html", users=user)
        except requests.exceptions.RequestException as e:
            message = f"Erro na solicitação: {e}"
            return render_template('error.html', message=message), 500

    else:
        return redirect(url_for('login'))


@app.route('/kit-colaborador', methods=['POST', 'GET'])
def kit_colaborador():
    if 'username' not in session:
        return redirect(url_for('login'))

    try:
        response = fetchbravas(ip)

        if response.status_code != 200:
            message = f"Erro na solicitação: {response.text}"
            return render_template('error.html', message=message), 500

        data = json.loads(response.text)
        users = data["config"]["users"]

        user_selections = request.form.getlist('user_selection[]')
        user_info_values = request.form.getlist('user_info[]')

        for selected, info in zip(user_selections, user_info_values):
            print(f'Selecionado: {selected}, User Info: {info}')

        return render_template("kit-colaborador.html", users=users)
    except requests.exceptions.RequestException as e:
        message = f"Erro na solicitação: {e}"
        return render_template('error.html', message=message), 500


@app.route('/kitedit', methods=['POST'])
def kitedit():
    data = request.get_json()

    for items in data:
        nome = items.get("nome")
        estado = items.get("estado")
        editkit(ip=ip, nome=nome, estado=estado)

    resposta = {'message': 'Processado com sucesso!'}
    return jsonify(resposta)


@app.route('/del-colaborador', methods=['POST'])
def delcolaborador():
    data = request.get_json()
    for items in data:
        nome = items.get("nome")
        delbravas(ip=ip, name=nome)

    resposta = {'message': 'Processado com sucesso!'}
    return jsonify(resposta)


@app.route('/user_info/<string:name>', methods=['POST', 'GET'])
def edit_user(name):
    form = EditColaborador()
    info_json = seluser(ip, name)
    info = json.loads(info_json.text)

    if form.validate_on_submit():
        cartao = form.cartao.data
        seg = form.seg.data
        ter = form.ter.data
        qua = form.qua.data
        qui = form.qui.data
        sex = form.sex.data
        sab = form.sab.data
        dom = form.dom.data
        sab = form.sab.data
        dom = form.dom.data
        cafe_manha = form.cafe_manha.data
        almoco = form.almoco.data
        cafe_pendura = form.cafe_pendura.data
        cafe_tarde = form.cafe_tarde.data
        janta = form.janta.data

        response = editbravas(ip, name, tag=cartao, seg=seg, ter=ter, qua=qua, qui=qui, sex=sex, sab=sab, dom=dom,
                              cafe_manha=cafe_manha,
                              almoco=almoco, cafe_pendura=cafe_pendura, cafe_tarde=cafe_tarde, janta=janta)
        if response.status_code == 200:
            flash('Colaborador atualizado com sucesso!', 'success')
        else:
            flash('Erro ao atualizar colaborador!', 'danger')

    return render_template('edit-colaborador.html', name=name, info=info, form=form)


@app.route('/del-geral', methods=['POST', 'GET'])
def delete():
    response = dellGeral(ip)
    return response
